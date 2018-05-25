#!/usr/bin/env python
# encoding: utf-8
import linecache
import json
import time
import openpyxl
new_book = openpyxl.Workbook()
new_sheet = new_book.create_sheet(title="Sheet1")

new_sheet.cell(row=1, column=1).value = "id"
new_sheet.cell(row=1, column=2, value="localid")
new_sheet.cell(row=1, column=3, value="server_time")
new_sheet.cell(row=1, column=4, value="receive_time")
new_sheet.cell(row=1, column=5, value="total_cost")
new_sheet.cell(row=1, column=6, value="server_to_receive_cost")
new_sheet.cell(row=1, column=7, value="send_to_server_cost")


# 获取txt文件的行数
def get_line_count2(filename):
    count = 0
    with open(filename, 'r', encoding='utf-8') as f:
        for file_line in f.readlines():
            file_line = file_line.strip()
            count += 1
    print(count)
    return count


# 获取指定行中的指定内容
def insert_result(content, row):
    # print(content)
    strs = content.split('=')[1]
    strs = strs.split(", receiveIMNum")[0]

    # print("str:", strs)
    jsons = json.loads(strs)
    id = str(jsons['id'])
    # print("id:", id)

    localids = jsons['localid']
    localid = localids.split('_')[0]
    # print("localid:", localid)

    server_time = str(jsons['time'])
    # print("server_time:", server_time)

    receive_time = content.split("time=")[1]
    # print("receive_time:", receive_time)

    total_cost = int(receive_time)-int(localid)
    server_to_receive_cost = int(receive_time)-int(server_time)
    send_to_server_cost = int(server_time)-int(localid)

    new_sheet.cell(row=row+1, column=1).value = id
    new_sheet.cell(row=row+1, column=2).value = localid
    new_sheet.cell(row=row+1, column=3).value = server_time
    new_sheet.cell(row=row+1, column=4).value = receive_time
    new_sheet.cell(row=row+1, column=5).value = total_cost
    new_sheet.cell(row=row+1, column=6).value = server_to_receive_cost
    new_sheet.cell(row=row+1, column=7).value = send_to_server_cost


# 读取指定文件
def get_rowdata(line_number, file):
    the_line = linecache.getline(file, line_number)
    # print(the_line)
    if line_number < 1:
        return ''
    for cur_line_number, line in enumerate(open(file, 'rU')):
        insert_result(line, cur_line_number+1)
        # print(cur_line_number)
        # print(line)
        # return line
    return ''


if __name__ == '__main__':
    file = "rec.log"
    date1=time.time()
    line_number = get_line_count2(file)
    the_line = get_rowdata(line_number, file)
    new_book.save("result.xls")

    date2=time.time()
    print(date2-date1)
